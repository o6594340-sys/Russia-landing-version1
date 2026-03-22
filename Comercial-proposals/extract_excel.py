import openpyxl
wb = openpyxl.load_workbook('Japan-rates-sample.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'--- {sheet_name} ---')
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if any(cell for cell in row):
            print(row)