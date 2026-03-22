"""
Генератор сметы в Excel (формат .xlsx).
Структура: шапка → размещение → день за днём → итого
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# Цветовая палитра
C_DARK    = '1C1C1E'
C_RED     = 'BC002D'
C_GOLD    = 'C5A028'
C_LGRAY   = 'F5F5F5'
C_MGRAY   = 'E8E8E8'
C_WHITE   = 'FFFFFF'
C_BORDER  = 'CCCCCC'


def _font(bold=False, size=9, color=C_DARK, italic=False, name='Arial'):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def _fill(color):
    return PatternFill(fill_type='solid', fgColor=color)


def _align(h='left', v='center', wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def _border_thin():
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_row(ws, row, col_values: dict, height=16):
    """Задаёт значения и стили строки."""
    for col, (value, style) in col_values.items():
        cell = ws.cell(row=row, column=col, value=value)
        if style:
            for attr, val in style.items():
                setattr(cell, attr, val)
    ws.row_dimensions[row].height = height


def create_excel(params: dict, services: list, content: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Смета Япония'

    # Ширины столбцов: A=дата B=blank C=сервис D=кол D=дни F=blank G=цена/ед H=итого I=коммент
    col_widths = {'A': 14, 'B': 3, 'C': 44, 'D': 9, 'E': 7, 'F': 3, 'G': 16, 'H': 16, 'I': 36}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ── ЗАГОЛОВОК ──────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = 'КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ — ЯПОНИЯ, ТОКИО'
    c.font = _font(bold=True, size=13, color=C_WHITE)
    c.fill = _fill(C_DARK)
    c.alignment = _align('center')
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = params['company_name']
    c.font = _font(bold=True, size=12, color=C_RED)
    c.fill = _fill(C_LGRAY)
    c.alignment = _align('center')
    ws.row_dimensions[row].height = 22
    row += 1

    # Детали запроса
    twn = params.get('twn', 0)
    sgl = params.get('sgl', 0)
    room_str = []
    if twn: room_str.append(f'{twn} Twin')
    if sgl: room_str.append(f'{sgl} Single')
    room_summary = ', '.join(room_str) if room_str else 'уточняется'
    row_data = [
        ('Количество человек', str(params['pax'])),
        ('Продолжительность', f"{params['days']} дней / {params['days'] - 1} ночей"),
        ('Даты', params.get('dates') or 'уточняются'),
        ('Тип размещения', f"{room_summary} | {params['hotel_level']}"),
        ('Тип мероприятия', params['event_type']),
    ]
    for label, val in row_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = _font(bold=True, size=9, color='666666')
        ws[f'A{row}'].fill = _fill(C_LGRAY)
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = val
        ws[f'B{row}'].font = _font(size=9)
        ws[f'B{row}'].fill = _fill(C_LGRAY)
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    # Важная пометка
    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = ('ВАЖНО! Все цены предварительные и подлежат подтверждению при бронировании. '
               'Бронирование не произведено. При изменении курса иены производится перерасчёт.')
    c.font = _font(italic=True, size=8, color='888888')
    c.alignment = _align(wrap=True)
    ws.row_dimensions[row].height = 24
    row += 1

    row += 1

    # ── ШАПКА ТАБЛИЦЫ ─────────────────────────────────────────────────────────
    headers = {
        1: 'Дата', 2: '', 3: 'Сервис / Позиция',
        4: 'Кол-во', 5: 'Дней', 6: '',
        7: 'Цена за ед. (USD)', 8: 'Итого (USD)', 9: 'Комментарий'
    }
    for col, h in headers.items():
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, size=9, color=C_WHITE)
        c.fill = _fill(C_DARK)
        c.alignment = _align('center', wrap=True)
        c.border = _border_thin()
    ws.row_dimensions[row].height = 22
    header_row = row
    row += 1

    # ── РАЗМЕЩЕНИЕ ────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = '  РАЗМЕЩЕНИЕ'
    c.font = _font(bold=True, size=10, color=C_WHITE)
    c.fill = _fill(C_RED)
    c.alignment = _align()
    ws.row_dimensions[row].height = 18
    row += 1

    nights = params['days'] - 1
    pax = params['pax']
    hotel_level = params['hotel_level']
    twn = params.get('twn', 0)
    sgl = params.get('sgl', 0)
    hotel_a_rate = float(params.get('hotel_a_rate') or 350)
    hotel_b_rate = float(params.get('hotel_b_rate') or 0)
    hotel_a_name = params.get('hotel_a_name') or 'Отель 1'
    hotel_b_name = params.get('hotel_b_name') or 'Отель 2'
    hotel_comment = 'Ориентировочная цена, подтверждается при бронировании'

    hotel_data_rows = []  # collect all hotel rows for grand total

    def _write_hotel_row(ws, row, service_label, qty, nts, rate, comment=''):
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.fill = _fill(C_LGRAY)
            c.font = _font(size=9)
            c.border = _border_thin()
        ws[f'C{row}'] = service_label
        ws[f'D{row}'] = qty
        ws[f'D{row}'].alignment = _align('center')
        ws[f'E{row}'] = nts
        ws[f'E{row}'].alignment = _align('center')
        ws[f'G{row}'] = rate
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].alignment = _align('right')
        ws[f'H{row}'] = f'=D{row}*E{row}*G{row}'
        ws[f'H{row}'].number_format = '#,##0.00'
        ws[f'H{row}'].alignment = _align('right')
        ws[f'H{row}'].font = _font(bold=True, size=9)
        if comment:
            ws[f'I{row}'] = comment
            ws[f'I{row}'].font = _font(italic=True, size=8, color='888888')
        ws.row_dimensions[row].height = 16

    # --- Вариант А ---
    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = f'  ВАРИАНТ А — {hotel_a_name}'
    c.font = _font(bold=True, size=9, color=C_WHITE)
    c.fill = _fill(C_RED)
    c.alignment = _align()
    ws.row_dimensions[row].height = 16
    hotel_a_header_row = row
    row += 1

    if twn > 0:
        _write_hotel_row(ws, row, f'Twin-номера ({hotel_level}, Токио)', twn, nights, hotel_a_rate,
                         hotel_comment if not sgl else '')
        hotel_data_rows.append(row)
        row += 1
    if sgl > 0:
        _write_hotel_row(ws, row, f'Single-номера ({hotel_level}, Токио)', sgl, nights, hotel_a_rate,
                         hotel_comment)
        hotel_data_rows.append(row)
        row += 1
    if twn == 0 and sgl == 0:
        # fallback: no room counts specified
        _write_hotel_row(ws, row, f'Номера ({hotel_level}, Токио)', 1, nights, hotel_a_rate, hotel_comment)
        hotel_data_rows.append(row)
        row += 1

    # --- Вариант Б (если задан) ---
    hotel_b_rows = []
    if hotel_b_rate > 0:
        row += 1  # blank separator
        ws.merge_cells(f'A{row}:I{row}')
        c = ws[f'A{row}']
        c.value = f'  ВАРИАНТ Б — {hotel_b_name}'
        c.font = _font(bold=True, size=9, color=C_WHITE)
        c.fill = _fill(C_RED)
        c.alignment = _align()
        ws.row_dimensions[row].height = 16
        row += 1

        if twn > 0:
            _write_hotel_row(ws, row, f'Twin-номера ({hotel_level}, Токио)', twn, nights, hotel_b_rate,
                             hotel_comment if not sgl else '')
            hotel_b_rows.append(row)
            row += 1
        if sgl > 0:
            _write_hotel_row(ws, row, f'Single-номера ({hotel_level}, Токио)', sgl, nights, hotel_b_rate,
                             hotel_comment)
            hotel_b_rows.append(row)
            row += 1
        if twn == 0 and sgl == 0:
            _write_hotel_row(ws, row, f'Номера ({hotel_level}, Токио)', 1, nights, hotel_b_rate, hotel_comment)
            hotel_b_rows.append(row)
            row += 1

    row += 1  # пустая строка

    # ── СЕРВИСЫ ПО ДНЯМ ───────────────────────────────────────────────────────
    service_rows = []
    alt = False  # чередование строк

    for item in services:
        if item['type'] == 'day_header':
            ws.merge_cells(f'A{row}:I{row}')
            c = ws[f'A{row}']
            c.value = f'  {item["day"]}'
            c.font = _font(bold=True, size=10, color=C_WHITE)
            c.fill = _fill(C_RED)
            c.alignment = _align()
            ws.row_dimensions[row].height = 18
            alt = False
            row += 1

        elif item['type'] == 'service':
            service_rows.append(row)
            fill_color = C_LGRAY if alt else C_WHITE
            alt = not alt

            for col in range(1, 10):
                c = ws.cell(row=row, column=col)
                c.fill = _fill(fill_color)
                c.font = _font(size=9)
                c.border = _border_thin()

            ws[f'C{row}'] = item['service']
            ws[f'D{row}'] = item['q']
            ws[f'D{row}'].alignment = _align('center')
            ws[f'E{row}'] = item['days']
            ws[f'E{row}'].alignment = _align('center')
            ws[f'G{row}'] = item['price_per_unit']
            ws[f'G{row}'].number_format = '#,##0.00'
            ws[f'G{row}'].alignment = _align('right')
            ws[f'H{row}'] = f'=D{row}*E{row}*G{row}'
            ws[f'H{row}'].number_format = '#,##0.00'
            ws[f'H{row}'].alignment = _align('right')
            ws[f'H{row}'].font = _font(bold=True, size=9)
            if item['comments']:
                ws[f'I{row}'] = item['comments']
                ws[f'I{row}'].font = _font(size=8, color='666666', italic=True)
            ws.row_dimensions[row].height = 16
            row += 1

    row += 1  # пустая строка

    # ── ИТОГО ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:G{row}')
    c = ws[f'A{row}']
    c.value = 'ИТОГО (наземная часть + размещение)'
    c.font = _font(bold=True, size=11, color=C_WHITE)
    c.fill = _fill(C_DARK)
    c.alignment = _align('right', indent=1)
    ws.row_dimensions[row].height = 24

    # Сумма: отель А + все сервисы
    all_h_rows = hotel_data_rows + service_rows
    sum_parts = '+'.join(f'H{r}' for r in all_h_rows)
    ws[f'H{row}'] = f'={sum_parts}'
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = _font(bold=True, size=13, color=C_WHITE)
    ws[f'H{row}'].fill = _fill(C_DARK)
    ws[f'H{row}'].alignment = _align('right')
    ws[f'I{row}'].fill = _fill(C_DARK)
    total_row = row
    row += 1

    row += 1

    # ── ЦЕНА НА ЧЕЛОВЕКА — Вариант А ──────────────────────────────────────────
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = f'Цена на человека — Вариант А (наземная часть + размещение, {pax} чел.)'
    ws[f'A{row}'].font = _font(bold=True, size=9, color='444444')
    ws[f'H{row}'] = f'=H{total_row}/{pax}'
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = _font(bold=True, size=11, color=C_RED)
    ws[f'H{row}'].alignment = _align('right')
    ws.row_dimensions[row].height = 18
    row += 1

    # ── ЦЕНА НА ЧЕЛОВЕКА — Вариант Б (если задан) ──────────────────────────────
    if hotel_b_rows:
        all_b_rows = hotel_b_rows + service_rows
        sum_b_parts = '+'.join(f'H{r}' for r in all_b_rows)

        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f'Цена на человека — Вариант Б (наземная часть + размещение, {pax} чел.)'
        ws[f'A{row}'].font = _font(bold=True, size=9, color='444444')
        ws[f'H{row}'] = f'=({sum_b_parts})/{pax}'
        ws[f'H{row}'].number_format = '#,##0.00'
        ws[f'H{row}'].font = _font(bold=True, size=11, color=C_RED)
        ws[f'H{row}'].alignment = _align('right')
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ── ФУТЕР ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:I{row}')
    c = ws[f'A{row}']
    c.value = ('Tozai Tours — DMC Japan  |  Коммерческое предложение действительно 14 дней  |  '
               'Все цены в USD  |  Международные перелёты не включены')
    c.font = _font(italic=True, size=8, color='999999')
    c.alignment = _align('center')

    # Закрепить строку заголовков
    ws.freeze_panes = f'C{header_row + 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
