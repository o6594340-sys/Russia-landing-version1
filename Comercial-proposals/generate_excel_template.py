import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Смета Япония"

# Headers
headers = ["Сервис", "Описание", "Кол-во", "Цена за ед. (USD)", "Итого (USD)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Sample data based on requirements (hotels, meals, programs)
data = [
    ["Отель", "Отель в Токио, 4* (двухместное размещение)", 10, 200, "=C2*D2"],
    ["Питание", "Завтрак + обед + ужин (японская кухня)", 10, 50, "=C3*D3"],
    ["Экскурсии", "Посещение храмов + лекции", 10, 100, "=C4*D4"],
    ["Трансфер", "Аэропорт - отель - аэропорт", 10, 30, "=C5*D5"],
    ["Дополнительно", "Концепция мероприятия", 1, 500, "=C6*D6"],
    ["Итого", "", "", "", "=SUM(E2:E5)+E6"]
]

for row, row_data in enumerate(data, 2):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        if col == 5 and row < 7:
            cell.value = value  # Formula
        cell.alignment = Alignment(horizontal='center')

# Borders
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row in ws.iter_rows(min_row=1, max_row=7):
    for cell in row:
        cell.border = thin_border

# Save
wb.save("Japan-Proposal-Template.xlsx")
print("Шаблон сметы создан: Japan-Proposal-Template.xlsx")