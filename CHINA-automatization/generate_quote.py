"""
generate_quote.py — генерация сметы в Excel для proposal DMC.
Использование: python generate_quote.py
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Цвета Cathay DMC ──────────────────────────────────────────────────────────
C_GREEN       = "B9DC9F"   # светло-зелёный (основной акцент)
C_DARK_GREEN  = "5A8A3C"   # тёмно-зелёный (заголовки разделов)
C_DARK        = "272721"   # почти чёрный
C_LIGHT_GREEN = "E8F5DA"   # очень светлый (подзаголовки)
C_TOTAL_BG    = "D6ECC4"   # светло-зелёный (строки итогов)
C_WHITE       = "FFFFFF"
C_GRAY        = "AAAAAA"   # строки qty=0

THIN        = Side(style="thin",   color="C8E0B0")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Стили ─────────────────────────────────────────────────────────────────────

def s_section(cell):
    cell.font      = Font(bold=True, color=C_WHITE, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_DARK_GREEN)
    cell.alignment = Alignment(vertical="center", indent=1)

def s_subhead(cell):
    cell.font      = Font(bold=True, color=C_DARK, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_LIGHT_GREEN)
    cell.alignment = Alignment(vertical="center", indent=1)

def s_total(cell, grand=False):
    if grand:
        cell.font = Font(bold=True, color=C_WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=C_DARK)
    else:
        cell.font = Font(bold=True, color=C_DARK, size=10)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
    cell.alignment     = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00'

def s_data(cell, gray=False):
    cell.font      = Font(color=C_GRAY if gray else C_DARK, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=False, indent=1)
    cell.border    = BORDER_THIN

def s_num(cell, gray=False):
    cell.font          = Font(color=C_GRAY if gray else C_DARK, size=10)
    cell.alignment     = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00'
    cell.border        = BORDER_THIN

def s_num_int(cell, gray=False):
    cell.font          = Font(color=C_GRAY if gray else C_DARK, size=10)
    cell.alignment     = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0'
    cell.border        = BORDER_THIN


# ── Данные proposal ───────────────────────────────────────────────────────────

P = {
    "code":       "MP-2610-SH-35PAX",
    "client":     "ТЕХНОНИКОЛЬ",
    "agency":     "RUSMICE",
    "pax_sgl":    30,
    "pax_dbl":    5,
    "pax_total":  35,
    "issued":     date(2026, 4, 14),
    "valid":      date(2026, 4, 21),
    "conference": True,
}

NIGHTS = 4   # ночей в Шанхае

HOTELS = [
    {
        "name": "Courtyard by Marriott Shanghai Central 4*",
        "tier": 1,
        "rooms": [
            {"desc": "Deluxe King Room, 36 sqm, incl. breakfast",    "price": 155, "type": "SGL"},
            {"desc": "Deluxe Twin Room, 36 sqm, incl. 2 breakfasts", "price": 170, "type": "TWN"},
        ],
    },
    {
        "name": "Shanghai Marriott Marquis City Centre 5*",
        "tier": 2,
        "rooms": [
            {"desc": "Deluxe King Room, 33 sqm, incl. breakfast",    "price": 245, "type": "SGL"},
            {"desc": "Deluxe Twin Room, 33 sqm, incl. 2 breakfasts", "price": 270, "type": "TWN"},
        ],
    },
    {
        "name": "The Portman Ritz-Carlton Shanghai 5*",
        "tier": 3,
        "rooms": [
            {"desc": "Deluxe King Room, 38 sqm, incl. breakfast",    "price": 380, "type": "SGL"},
            {"desc": "Deluxe Twin Room, 38 sqm, incl. 2 breakfasts", "price": 410, "type": "TWN"},
        ],
    },
]

DAYS = [
    {
        "label": "DAY 1 — 13 Apr 2026 — Shanghai (Arrival)",
        "date":  "13 Apr", "city": "Shanghai",
        "items": [
            ("Transfer", "Coach 40 pax",     "Airport → Hotel, arrival transfer", "vehicle", 320, 1),
            ("Guide",    "RU Speaking Guide","Airport meet & greet, half day",    "shift",   190, 1),
        ],
    },
    {
        "label": "DAY 2 — 14 Apr 2026 — Shanghai",
        "date":  "14 Apr", "city": "Shanghai",
        "items": [
            ("Guide",    "RU Speaking Guide",       "Full day service",                        "day",  350,  1),
            ("Transfer", "Coach 40 pax",            "City transfers, full day",                "day",  420,  1),
            ("F&B",      "Lunch — The People's Park","Lunch (water, tea/coffee)",              "pax",   45, 35),
            ("Activity", "Shanghai History Museum", "Entrance tickets",                        "pax",   18, 35),
            ("Activity", "Shanghai Tower — Obs.118F","Entrance tickets, 118th floor",         "pax",   35, 35),
            ("F&B",      "Dinner — Lost Heaven",    "Dinner (2 glasses wine, tea/coffee)",     "pax",   88, 35),
        ],
    },
    {
        "label": "DAY 3 — 15 Apr 2026 — Suzhou",
        "date":  "15 Apr", "city": "Shanghai / Suzhou",
        "items": [
            ("Guide",    "RU Speaking Guide",        "Full day service",                       "day",  350,  1),
            ("Transfer", "Coach 40 pax",             "Shanghai → Suzhou → Shanghai",           "day",  580,  1),
            ("Activity", "Suzhou Museum",            "Entrance tickets",                       "pax",   12, 35),
            ("Activity", "Master of Nets Garden",    "Entrance tickets",                       "pax",   20, 35),
            ("F&B",      "Lunch — Restaurant Mulan", "Lunch (water, tea/coffee)",              "pax",   42, 35),
            ("Activity", "Seal Engraving Masterclass","Incl. materials, Xiling master",       "pax",   55, 35),
            ("F&B",      "Dinner — MeiLongZhen",     "Dinner (2 glasses wine, tea/coffee)",    "pax",   75, 35),
        ],
    },
    {
        "label": "DAY 4 — 16 Apr 2026 — Shanghai",
        "date":  "16 Apr", "city": "Shanghai",
        "items": [
            ("Guide",    "RU Speaking Guide",  "Full day service",                            "day",  350,  1),
            ("Transfer", "Coach 40 pax",       "City transfers, full day",                    "day",  420,  1),
            ("F&B",      "Lunch — Il Teatro",  "Lunch (water, tea/coffee)",                   "pax",   52, 35),
            ("Activity", "French Concession",  "Licensed guide, 2 hrs",                       "group",180,  1),
            ("F&B",      "Dinner — Gongyan",   "Dinner + show (2 glasses wine)",              "pax",  118, 35),
        ],
    },
    {
        "label": "DAY 5 — 17 Apr 2026 — Departure",
        "date":  "17 Apr", "city": "Shanghai",
        "items": [
            ("Guide",    "RU Speaking Guide","Airport transfer, half day",                    "shift",190,  1),
            ("Transfer", "Coach 40 pax",    "Hotel → Airport, departure transfer",            "vehicle",320,1),
        ],
    },
]

CONFERENCE = {
    "date": "14 Apr", "city": "Shanghai",
    "items": [
        ("Conference", "Meeting Room — Courtyard Marriott","Full day rental, up to 40 pax",          "day", 850,  1),
        ("Conference", "AV Equipment",                     "Projector, screen, microphones, laptop", "set", 380,  1),
        ("Conference", "Lunch at Hotel — Buffet",          "Buffet lunch, water & soft drinks incl.","pax",  62, 35),
    ],
}


# ── Генератор ─────────────────────────────────────────────────────────────────

def generate(filename="Quote_sample.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    widths = [2, 12, 20, 12, 28, 44, 12, 12, 7, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = [1]

    def R():     return row[0]
    def inc(n=1): row[0] += n

    def rh(h):   ws.row_dimensions[R()].height = h

    # ── вспомогательные функции ───────────────────────────────────────────

    def merge_write(col, text, style_fn, height=18):
        rh(height)
        ws.merge_cells(f"{get_column_letter(col)}{R()}:J{R()}")
        c = ws.cell(R(), col, text)
        style_fn(c)
        inc()

    def write_col_headers():
        rh(18)
        hdrs = ["DATE","CITY","TYPE","SERVICE","DESCRIPTION","UNIT","PRICE (USD)","QTY","TOTAL (USD)"]
        for col, h in enumerate(hdrs, 2):
            c = ws.cell(R(), col, h)
            c.font      = Font(bold=True, color=C_WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=C_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = BORDER_THIN
        inc()

    def write_row(date_s, city, type_, service, desc, unit, price, qty):
        """Записывает строку данных. Возвращает total (число)."""
        total = round(price * qty, 2)
        gray  = (qty == 0)
        rh(15)
        for col, val in enumerate([date_s, city, type_, service, desc, unit], 2):
            s_data(ws.cell(R(), col, val), gray)
        s_num(ws.cell(R(), 8, price), gray)
        s_num_int(ws.cell(R(), 9, qty), gray)
        tc = ws.cell(R(), 10, total)
        s_num(tc, gray)
        inc()
        return total

    def write_total(label, section_total):
        rh(18)
        ws.merge_cells(f"B{R()}:I{R()}")
        c = ws.cell(R(), 2, label)
        s_total(c)
        tc = ws.cell(R(), 10, round(section_total, 2))
        s_total(tc)
        inc(2)
        return section_total

    # ════════════════════════════════════════════════════════════════════
    # ШАПКА
    # ════════════════════════════════════════════════════════════════════
    rh(55)
    c = ws.cell(R(), 2, "CATHAY DMC\nFresh Perspective on China\nwww.cathaydmc.com/ru")
    c.font      = Font(bold=True, size=13, color=C_DARK)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(f"B{R()}:J{R()}")
    inc()

    merge_write(2, "PRELIMINARY PROJECT QUOTATION",
                lambda c: setattr(c, 'font', Font(bold=True, size=12, color=C_DARK_GREEN)), 20)
    inc()

    meta = [
        ("Project code:",     P["code"]),
        ("Client:",           P["client"]),
        ("Agency:",           P["agency"]),
        ("Number of guests:", f"{P['pax_total']} pax  ({P['pax_sgl']} SGL + {P['pax_dbl']} DBL)"),
        ("Date issued:",      P["issued"].strftime("%d %b %Y")),
        ("Date valid:",       P["valid"].strftime("%d %b %Y")),
    ]
    for label, value in meta:
        rh(15)
        ws.cell(R(), 2, label).font = Font(bold=True, size=10)
        ws.merge_cells(f"C{R()}:J{R()}")
        ws.cell(R(), 3, value).font = Font(size=10)
        inc()
    inc()

    write_col_headers()

    # ════════════════════════════════════════════════════════════════════
    # ACCOMMODATION
    # ════════════════════════════════════════════════════════════════════
    merge_write(2, "ACCOMMODATION  —  13–16 Apr 2026  —  Shanghai", s_section, 20)
    accom_total = 0.0

    for hotel in HOTELS:
        merge_write(2, "  " + hotel["name"], s_subhead, 16)
        for rm in hotel["rooms"]:
            if hotel["tier"] == 1:
                qty = P["pax_sgl"] * NIGHTS if rm["type"] == "SGL" else P["pax_dbl"] * NIGHTS
            else:
                qty = 0
            accom_total += write_row(
                "13–16 Apr", "Shanghai", "Hotel",
                hotel["name"], rm["desc"], "room*night",
                rm["price"], qty
            )
        inc()   # пустая строка между отелями

    write_total("TOTAL ACCOMMODATION", accom_total)

    # ════════════════════════════════════════════════════════════════════
    # EVENTS: трансферы, гиды, питание, активности
    # ════════════════════════════════════════════════════════════════════
    merge_write(2, "TRANSFERS, GUIDES, F&B & ACTIVITIES", s_section, 20)
    events_total = 0.0

    for day in DAYS:
        merge_write(2, "  " + day["label"], s_subhead, 16)
        for type_, service, desc, unit, price, qty in day["items"]:
            events_total += write_row(day["date"], day["city"],
                                      type_, service, desc, unit, price, qty)
        inc()

    write_total("TOTAL TRANSFERS, GUIDES, F&B & ACTIVITIES", events_total)

    # ════════════════════════════════════════════════════════════════════
    # CONFERENCE (опционально)
    # ════════════════════════════════════════════════════════════════════
    conf_total = 0.0
    if P["conference"]:
        merge_write(2, "CONFERENCE SERVICES", s_section, 20)
        for type_, service, desc, unit, price, qty in CONFERENCE["items"]:
            conf_total += write_row(CONFERENCE["date"], CONFERENCE["city"],
                                    type_, service, desc, unit, price, qty)
        write_total("TOTAL CONFERENCE SERVICES", conf_total)

    # ════════════════════════════════════════════════════════════════════
    # GRAND TOTAL
    # ════════════════════════════════════════════════════════════════════
    inc()
    grand = round(accom_total + events_total + conf_total, 2)

    rh(22)
    ws.merge_cells(f"B{R()}:I{R()}")
    c = ws.cell(R(), 2, "GRAND TOTAL (USD)")
    s_total(c, grand=True)
    tc = ws.cell(R(), 10, grand)
    s_total(tc, grand=True)
    inc()

    rh(22)
    ws.merge_cells(f"B{R()}:I{R()}")
    c = ws.cell(R(), 2, f"PER PERSON (USD)  —  based on {P['pax_total']} pax")
    s_total(c, grand=True)
    tc = ws.cell(R(), 10, round(grand / P["pax_total"], 2))
    s_total(tc, grand=True)

    ws.freeze_panes = "B12"

    wb.save(filename)
    print(f"Готово: {filename}")
    print(f"  Размещение:        ${accom_total:,.2f}")
    print(f"  Трансферы/гиды/FB: ${events_total:,.2f}")
    print(f"  Конференция:       ${conf_total:,.2f}")
    print(f"  ИТОГО:             ${grand:,.2f}")
    print(f"  На человека:       ${grand/P['pax_total']:,.2f}")


if __name__ == "__main__":
    generate()
